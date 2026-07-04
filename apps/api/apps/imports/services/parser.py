from __future__ import annotations

import csv
from datetime import datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO
from pathlib import Path

from openpyxl import load_workbook


REQUIRED_COLUMNS = {"medicine name", "quantity"}


class ImportParseError(ValueError):
    pass


def normalize_header(value: str) -> str:
    return " ".join((value or "").strip().lower().replace("_", " ").split())


def parse_decimal(value):
    if value in (None, ""):
        return None
    try:
        return Decimal(str(value))
    except InvalidOperation as exc:
        raise ValueError("Invalid decimal value.") from exc


def parse_int(value):
    if value in (None, ""):
        return None
    try:
        return int(value)
    except (TypeError, ValueError) as exc:
        raise ValueError("Invalid integer value.") from exc


def parse_date(value):
    if value in (None, ""):
        return None
    if hasattr(value, "date"):
        return value.date()
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError("Invalid date value.")


def read_rows(uploaded_file):
    suffix = Path(uploaded_file.name).suffix.lower()
    data = uploaded_file.read()
    if suffix == ".csv":
        text = data.decode("utf-8-sig")
        reader = csv.DictReader(StringIO(text))
        headers = [normalize_header(h) for h in (reader.fieldnames or [])]
        if not REQUIRED_COLUMNS.issubset(headers):
            raise ImportParseError("Missing required columns: Medicine name and Quantity.")
        return [{normalize_header(key): value for key, value in row.items()} for row in reader]
    if suffix == ".xlsx":
        workbook = load_workbook(BytesIO(data), data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            raise ImportParseError("Import file is empty.")
        headers = [normalize_header(str(cell or "")) for cell in rows[0]]
        if not REQUIRED_COLUMNS.issubset(headers):
            raise ImportParseError("Missing required columns: Medicine name and Quantity.")
        parsed = []
        for values in rows[1:]:
            parsed.append({headers[idx]: values[idx] if idx < len(values) else None for idx in range(len(headers))})
        return parsed
    raise ImportParseError("Unsupported file type. Upload CSV or XLSX.")


def normalize_row(raw: dict) -> dict:
    return {
        "medicine_name": raw.get("medicine name", ""),
        "quantity": parse_int(raw.get("quantity")),
        "generic_name": raw.get("generic name", ""),
        "strength": raw.get("strength", ""),
        "form": raw.get("form", ""),
        "batch_number": raw.get("batch number", "") or "",
        "expiry_date": parse_date(raw.get("expiry date")),
        "supplier_name": raw.get("supplier", "") or raw.get("supplier name", "") or "",
        "purchase_cost": parse_decimal(raw.get("purchase cost")),
        "selling_price": parse_decimal(raw.get("selling price")),
        "low_stock_threshold": parse_int(raw.get("low stock threshold")) or 5,
    }

